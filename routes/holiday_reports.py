from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required
from services.report_service import compute_payment_sheet, compute_present_status, process_holiday_duty, get_holiday_records
from models.holiday import Holiday, HolidayDutyRecord
from extensions import db
from datetime import datetime
from collections import defaultdict
import io
import json
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

holiday_reports_bp = Blueprint('holiday_reports', __name__)

@holiday_reports_bp.route('/reports')
@login_required
def reports_page():
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('reports.html', today=today)

@holiday_reports_bp.route('/present_status')
@login_required
def present_status_page():
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('present_filter.html', today=today)

@holiday_reports_bp.route('/api/reports/payment_sheet', methods=['POST'])
@login_required
def api_payment_sheet():
    data = request.get_json(silent=True) or {}
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    if not for_date:
        return jsonify({'error': 'date is required (YYYY-MM-DD)'}), 400
    try:
        # Check if this date is a holiday and has been processed
        holiday = Holiday.query.filter_by(holiday_date=datetime.strptime(for_date, '%Y-%m-%d').date()).first()
        if holiday and holiday.processed_at:
            rows = get_holiday_records(holiday.id, section, sub_section, category)
        else:
            # If not a holiday or not processed, we don't return regular data in this module
            # as per user request: "Just holiday data will come"
            rows = []
        return jsonify({'date': for_date, 'section': section, 'sub_section': sub_section, 'category': category, 'rows': rows})
    except Exception as e:
        print('payment_sheet error:', e)
        return jsonify({'error': f'failed to compute: {str(e)}'}), 500

@holiday_reports_bp.route('/api/holidays', methods=['GET', 'POST'])
@login_required
def api_holidays():
    if request.method == 'POST':
        data = request.get_json()
        h_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        h_name = data.get('name', 'Holiday')
        
        # Check if already exists
        existing = Holiday.query.filter_by(holiday_date=h_date).first()
        if existing:
            return jsonify({'error': 'Holiday already defined for this date'}), 400
            
        new_holiday = Holiday(holiday_date=h_date, holiday_name=h_name)
        db.session.add(new_holiday)
        db.session.commit()
        return jsonify({'id': new_holiday.id, 'date': str(new_holiday.holiday_date), 'name': new_holiday.holiday_name})
    
    holidays = Holiday.query.order_by(Holiday.holiday_date.desc()).all()
    return jsonify([{
        'id': h.id, 
        'date': str(h.holiday_date), 
        'name': h.holiday_name, 
        'status': h.status,
        'processed_at': h.processed_at.strftime('%Y-%m-%d %H:%M') if h.processed_at else None
    } for h in holidays])

@holiday_reports_bp.route('/api/holidays/<int:h_id>', methods=['DELETE'])
@login_required
def api_delete_holiday(h_id):
    from flask_login import current_user
    holiday = Holiday.query.get_or_404(h_id)
    
    # Allow deletion if it's draft OR if the user is an Admin
    is_admin = getattr(current_user, 'role', 'User') == 'Admin'
    
    if holiday.status == 'finalized' and not is_admin:
        return jsonify({'error': 'Only Admins can delete a finalized holiday'}), 400
        
    db.session.delete(holiday)
    db.session.commit()
    return jsonify({'message': 'deleted'})

@holiday_reports_bp.route('/api/holidays/<int:h_id>/process', methods=['POST'])
@login_required
def api_process_holiday(h_id):
    try:
        count = process_holiday_duty(h_id)
        return jsonify({'message': f'Processed {count} records'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@holiday_reports_bp.route('/api/holidays/<int:h_id>/finalize', methods=['POST'])
@login_required
def api_finalize_holiday(h_id):
    holiday = Holiday.query.get_or_404(h_id)
    holiday.status = 'finalized'
    db.session.commit()
    return jsonify({'message': 'finalized'})

@holiday_reports_bp.route('/api/reports/present_status', methods=['POST'])
@login_required
def api_present_status():
    data = request.get_json(silent=True) or {}
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    status = data.get('status', 'all')
    if not for_date:
        return jsonify({'error': 'date is required'}), 400
    try:
        rows = compute_present_status(for_date, section, sub_section, category, status)
        return jsonify({'date': for_date, 'rows': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@holiday_reports_bp.route('/reports/payment_sheet/pdf', methods=['POST'])
@login_required
def payment_sheet_pdf():
    if request.is_json:
        data = request.get_json(silent=True) or {}
    else:
        form_data = request.form.get('data')
        data = json.loads(form_data) if form_data else {}
    
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    if not for_date:
        return jsonify({'error': 'date is required'}), 400
    
    # Use snapshot if holiday exists and is processed
    holiday = Holiday.query.filter_by(holiday_date=datetime.strptime(for_date, '%Y-%m-%d').date()).first()
    if holiday and holiday.processed_at:
        rows = get_holiday_records(holiday.id, section, sub_section, category)
    else:
        rows = []
    
    grouped_rows = defaultdict(list)
    for r in rows:
        sec_name = (r.get('section') or 'All').strip()
        grouped_rows[sec_name].append(r)

    html_content = render_template(
        'payment_sheet_pdf.html',
        for_date=for_date,
        grouped_rows=grouped_rows,
        holiday=holiday,
        datetime=datetime
    )

    pdf_buffer = io.BytesIO()
    HTML(string=html_content, base_url=request.url_root).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)
    return send_file(pdf_buffer, as_attachment=False, mimetype='application/pdf')

@holiday_reports_bp.route('/reports/present_status/pdf', methods=['POST'])
@login_required
def present_status_pdf():
    if request.is_json:
        data = request.get_json(silent=True) or {}
    else:
        form_data = request.form.get('data')
        data = json.loads(form_data) if form_data else {}
    
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    status = data.get('status', 'all')
    if not for_date:
        return jsonify({'error': 'date is required'}), 400
    
    rows = compute_present_status(for_date, section, sub_section, category, status)
    
    grouped_rows = defaultdict(list)
    for r in rows:
        sec_name = (r.get('section') or 'All').strip()
        grouped_rows[sec_name].append(r)

    html_content = render_template(
        'present_status.html',
        for_date=for_date,
        grouped_rows=grouped_rows,
        datetime=datetime
    )

    pdf_buffer = io.BytesIO()
    HTML(string=html_content, base_url=request.url_root).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)
    return send_file(pdf_buffer, as_attachment=False, mimetype='application/pdf')

@holiday_reports_bp.route('/reports/payment_sheet/excel', methods=['POST'])
@login_required
def payment_sheet_excel():
    data = request.get_json(silent=True) or {}
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    if not for_date:
        return jsonify({'error': 'date is required'}), 400
    
    # Use snapshot if holiday exists and is processed
    holiday = Holiday.query.filter_by(holiday_date=datetime.strptime(for_date, '%Y-%m-%d').date()).first()
    if holiday and holiday.processed_at:
        rows = get_holiday_records(holiday.id, section, sub_section, category)
    else:
        rows = []

    wb = Workbook()
    ws = wb.active
    ws.title = 'Payment Sheet'
    headers = ['SL', 'ID', 'Name', 'Designation', 'Section', 'Gross', 'Basic', 'In Time', 'Out Time', 'Amount', 'Signature']
    ws.append(headers)
    for r in rows:
        ws.append([
            r['sl'], r['id'], r['name'], r['designation'], r['section'], r['gross'], round(r['basic'], 0), r['in_time'], r['out_time'], r['amount'], ''
        ])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue()), as_attachment=True,
                     download_name=f'payment_sheet_{for_date}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
