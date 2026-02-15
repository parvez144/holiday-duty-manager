from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required
from services.employee_service import get_employees, get_distinct_sections, get_distinct_categories
from services.attendance_service import get_attendance_for_date
from datetime import datetime
from collections import defaultdict
import io
import json
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/reports')
@login_required
def reports_page():
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('reports.html', today=today)


@reports_bp.route('/present_status')
@login_required
def present_status_page():
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('present_filter.html', today=today)


@reports_bp.route('/api/reports/sections')
@login_required
def api_sections():
    """Return distinct section list from employees."""
    return jsonify(get_distinct_sections())


@reports_bp.route('/api/reports/sub_sections')
@login_required
def api_sub_sections():
    """Return distinct sub_section list from employees, optionally filtered by section."""
    from services.employee_service import get_distinct_sub_sections
    section = request.args.get('section')
    return jsonify(get_distinct_sub_sections(section=section))


@reports_bp.route('/api/reports/categories')
@login_required
def api_categories():
    """Return distinct category list from employees."""
    return jsonify(get_distinct_categories())


def _compute_payment_sheet(for_date: str, section: str | None, sub_section: str | None, category: str | None):
    """Compute payment sheet rows for a given date and optional section/sub_section.
    
    Uses employee data from employee_service and 
    attendance logs from attendance_service.
    """
    # 1. Fetch Employees
    employees = get_employees(section=section, sub_section=sub_section, category=category)
    if not employees:
        return []

    # 2. Fetch Attendance
    # We can optimize by passing IDs, or just fetch all for the day if the list is huge.
    # Given the implementation, passing IDs is safer for filtering.
    emp_ids = [str(e['Emp_Id']) for e in employees]
    attendance_data = get_attendance_for_date(for_date, emp_ids)

    rows = []
    serial = 1
    
    for emp in employees:
        emp_id = str(emp['Emp_Id'])
        
        # Skip Security personnel as they don't get holiday payment
        sec = (emp.get('Section') or '').strip().lower()
        sub_sec = (emp.get('Sub_Section') or '').strip().lower()
        if sec == 'security' or sub_sec == 'security':
            continue

        stats = attendance_data.get(emp_id)
        
        if not stats:
            continue # Skip if person has NO punches at all on this day

        # Partitioned Times from Service
        in_dt = stats.get('in_time')
        out_dt = stats.get('out_time')

        # Start Time Rule: Cleaner @ 7:30 AM, Others @ 8:00 AM
        if in_dt:
            is_cleaner = (emp.get('Sub_Section') or '').strip().lower() == 'cleaner'
            start_h, start_m = (7, 30) if is_cleaner else (8, 0)
            start_limit = in_dt.replace(hour=start_h, minute=start_m, second=0, microsecond=0)
            eff_in = max(in_dt, start_limit)
            disp_in = eff_in.strftime('%H:%M')
        else:
            eff_in = None
            disp_in = "Missing"

        # 30-Minute Rounding Down for Out-Time
        if out_dt:
            eff_out = out_dt.replace(minute=(out_dt.minute // 30) * 30, second=0, microsecond=0)
            disp_out = eff_out.strftime('%H:%M')
        else:
            eff_out = None
            disp_out = "Missing"

        # Duration Calculation based on effective times
        if eff_in and eff_out:
            duration = eff_out - eff_in
            raw_hours = duration.total_seconds() / 3600.0
            # Deduct 1 hour for lunch ONLY if working 6 hours or more
            deduction = 1.0 if raw_hours >= 6.0 else 0.0
            work_hours = max(0, raw_hours - deduction)
        else:
            work_hours = 0
        
        # Salary calculations
        gross_salary = float(emp.get('Gross_Salary') or 0)
        # Basic = (Gross - allowances)/1.5
        basic_salary = (gross_salary - 2450) / 1.5
        daily_basic = basic_salary / 30.0
        ot_rate_unit = (basic_salary / 208.0) * 2.0

        category = (emp.get('Category') or '').strip().lower()
        
        # Robust check to match 'worker', 'workers', 'factory worker' etc.
        if 'worker' in category:
            # Workers: Entire duration as OT
            ot_hours = work_hours
            ot_rate = ot_rate_unit
            amount = ot_hours * ot_rate
        else:
            # Staff and others: One day basic money
            ot_hours = 0
            ot_rate = ot_rate_unit
            amount = daily_basic

        # Force 0 amount if any punch is missing/invalid
        # User Rule Update: For Staff, 1 punch is enough for amount. 
        # For Workers, 2 punches are required as OT depends on duration.
        if 'Missing' in disp_in or 'Missing' in disp_out:
            if 'worker' in category:
                amount = 0
                ot_hours = 0
                work_hours = 0
            else:
                # Non-worker (Staff): Single punch is enough to get daily basic.
                # However, if both happen to be Missing, amount is 0.
                if 'Missing' in disp_in and 'Missing' in disp_out:
                    amount = 0
                else:
                    # Amount is already daily_basic from line 118
                    pass
                ot_hours = 0
                work_hours = 0

        rows.append({
            'sl': serial,
            'id': emp_id,
            'name': emp['Emp_Name'].title() if emp['Emp_Name'] else '',
            'designation': emp['Designation'].title() if emp['Designation'] else '',
            'sub_section': emp['Sub_Section'].title() if emp['Sub_Section'] else '',
            'section': emp['Section'].title() if emp['Section'] else '',
            'category': emp.get('Category', ''),
            'gross': gross_salary,
            'basic': round(basic_salary, 0),
            'in_time': disp_in,
            'out_time': disp_out,
            'hour': round(work_hours, 2),
            'ot': round(ot_hours, 2),
            'ot_rate': round(ot_rate, 2),
            'amount': round(amount, 0),
            'remarks': '',
            'signature': ''
        })
        serial += 1

    return rows


def _compute_present_status(for_date: str, section: str | None, sub_section: str | None, category: str | None):
    """Compute present status rows (attendance only)."""
    employees = get_employees(section=section, sub_section=sub_section, category=category)
    if not employees:
        return []

    emp_ids = [str(e['Emp_Id']) for e in employees]
    attendance_data = get_attendance_for_date(for_date, emp_ids)

    rows = []
    serial = 1
    
    for emp in employees:
        emp_id = str(emp['Emp_Id'])
        stats = attendance_data.get(emp_id)
        
        # In this report, only skip if NO punches at all on this day
        if not stats or (not stats.get('in_time') and not stats.get('out_time')):
            continue

        in_dt = stats.get('in_time')
        out_dt = stats.get('out_time')

        # Show RAW Times directly from database
        disp_in = in_dt.strftime('%H:%M') if in_dt else "Missing"
        disp_out = out_dt.strftime('%H:%M') if out_dt else " Missing"

        rows.append({
            'sl': serial,
            'id': emp_id,
            'name': emp['Emp_Name'].title() if emp['Emp_Name'] else '',
            'designation': emp['Designation'].title() if emp['Designation'] else '',
            'sub_section': emp['Sub_Section'].title() if emp['Sub_Section'] else '',
            'in_time': disp_in,
            'out_time': disp_out,
            'remarks': ''
        })
        serial += 1

    return rows


@reports_bp.route('/api/reports/payment_sheet', methods=['POST'])
def api_payment_sheet():
    data = request.get_json(silent=True) or {}
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    if not for_date:
        return jsonify({'error': 'date is required (YYYY-MM-DD)'}), 400
    try:
        rows = _compute_payment_sheet(for_date, section, sub_section, category)
        return jsonify({'date': for_date, 'section': section, 'sub_section': sub_section, 'category': category, 'rows': rows})
    except Exception as e:
        print('payment_sheet error:', e)
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'failed to compute: {str(e)}'}), 500


@reports_bp.route('/reports/payment_sheet/pdf', methods=['POST'])
def payment_sheet_pdf():
    # Accept both JSON and form data
    if request.is_json:
        data = request.get_json(silent=True) or {}
    else:
        # Handle form data
        form_data = request.form.get('data')
        if form_data:
            data = json.loads(form_data)
        else:
            data = {}
    
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    if not for_date:
        return jsonify({'error': 'date is required'}), 400
    
    rows = _compute_payment_sheet(for_date, section, sub_section, category)
    
    grouped_rows = defaultdict(list)
    for r in rows:
        sec_name = (r.get('section') or 'All').strip()
        grouped_rows[sec_name].append(r)

    html_content = render_template(
        'payment_sheet_pdf.html',
        for_date=for_date,
        grouped_rows=grouped_rows,
        datetime=datetime
    )

    pdf_buffer = io.BytesIO()
    HTML(string=html_content, base_url=request.url_root).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)

    return send_file(pdf_buffer, as_attachment=False, mimetype='application/pdf')


@reports_bp.route('/api/reports/present_status', methods=['POST'])
def api_present_status():
    data = request.get_json(silent=True) or {}
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    if not for_date:
        return jsonify({'error': 'date is required'}), 400
    try:
        rows = _compute_present_status(for_date, section, sub_section, category)
        return jsonify({'date': for_date, 'rows': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/reports/present_status/pdf', methods=['POST'])
def present_status_pdf():
    # Handle both JSON and form data
    if request.is_json:
        data = request.get_json(silent=True) or {}
    else:
        form_data = request.form.get('data')
        data = json.loads(form_data) if form_data else {}
    
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    
    if not for_date:
        return jsonify({'error': 'date is required'}), 400
    
    rows = _compute_present_status(for_date, section, sub_section, category)
    
    grouped_rows = defaultdict(list)
    for r in rows:
        sec_name = (r.get('section') or 'All').strip()
        grouped_rows[sec_name].append(r)

    html_content = render_template(
        'present_status.html',
        for_date=for_date,
        grouped_rows=grouped_rows,
        datetime=datetime
    )

    pdf_buffer = io.BytesIO()
    HTML(string=html_content, base_url=request.url_root).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)

    return send_file(pdf_buffer, as_attachment=False, mimetype='application/pdf')


@reports_bp.route('/reports/payment_sheet/excel', methods=['POST'])
def payment_sheet_excel():
    data = request.get_json(silent=True) or {}
    for_date = data.get('date')
    section = data.get('section')
    sub_section = data.get('sub_section')
    category = data.get('category')
    if not for_date:
        return jsonify({'error': 'date is required'}), 400
    
    rows = _compute_payment_sheet(for_date, section, sub_section, category)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Payment Sheet'
    headers = ['SL', 'ID', 'Name', 'Designation', 'Section', 'Gross', 'Basic', 'In Time', 'Out Time', 'Amount', 'Signature']
    ws.append(headers)
    for r in rows:
        ws.append([
            r['sl'], r['id'], r['name'], r['designation'], r['section'], r['gross'], round(r['basic'], 0), r['in_time'], r['out_time'], r['amount'], ''
        ])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue()), as_attachment=True,
                     download_name=f'payment_sheet_{for_date}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
